from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
from calendar import monthrange
from openpyxl.utils import get_column_letter

def generate_gantt_excel(tasks: list, filename: str = "gantt_chart.xlsx") -> None:
    """
    Creates a Gantt chart in Excel format, showing only workdays
    
    Args:
        tasks (list): List of dictionaries containing task information
            Each dict should have: name, start_date, end_date
        filename (str): Name of the output Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    task_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title row first
    title_cell = ws.cell(row=1, column=1, value="Project Timeline")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Get project date range
    start_dates = [datetime.strptime(task['start_date'].replace('-00', '-01'), '%Y-%m-%d') for task in tasks]
    end_dates = [datetime.strptime(task['end_date'].replace('-00', '-01'), '%Y-%m-%d') for task in tasks]
    project_start = min(start_dates)
    project_end = max(end_dates)
    
    # Create date headers
    current_date = project_start
    date_columns = []
    col = 2  # Start from column B (A is for task names)
    
    # Format task names column
    ws.column_dimensions['A'].width = 40
    
    # Add month headers (Row 2) and day headers (Row 3)
    current_month = None
    month_start_col = col
    
    while current_date <= project_end:
        # Skip weekends
        if current_date.weekday() >= 5:
            current_date += timedelta(days=1)
            continue
        
        # Set column width
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 3
        
        # Handle month headers
        month_str = current_date.strftime('%B %Y')
        if current_month != month_str:
            if current_month is not None and col > month_start_col:
                # Merge previous month cells
                merge_range = f'{get_column_letter(month_start_col)}2:{get_column_letter(col-1)}2'
                ws.merge_cells(merge_range)
                ws.cell(row=2, column=month_start_col).fill = header_fill
                ws.cell(row=2, column=month_start_col).font = header_font
            
            current_month = month_str
            month_start_col = col
            ws.cell(row=2, column=col, value=current_month)
        
        # Add day number
        ws.cell(row=3, column=col, value=current_date.day)
        date_columns.append((col, current_date))
        
        col += 1
        current_date += timedelta(days=1)
    
    # Merge the last month
    if col > month_start_col:
        merge_range = f'{get_column_letter(month_start_col)}2:{get_column_letter(col-1)}2'
        ws.merge_cells(merge_range)
        ws.cell(row=2, column=month_start_col).fill = header_fill
        ws.cell(row=2, column=month_start_col).font = header_font
    
    # Merge title cell
    title_merge_range = f'A1:{get_column_letter(col-1)}1'
    ws.merge_cells(title_merge_range)
    
    # Add "Tasks" header
    ws.cell(row=2, column=1, value="Tasks")
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).font = header_font
    ws.cell(row=3, column=1, value="")  # Empty cell for day numbers row
    
    # Add tasks
    for row, task in enumerate(tasks, start=4):
        # Add task name
        ws.cell(row=row, column=1, value=task['name'])
        
        # Add task bar
        task_start = datetime.strptime(task['start_date'].replace('-00', '-01'), '%Y-%m-%d')
        task_end = datetime.strptime(task['end_date'].replace('-00', '-01'), '%Y-%m-%d')
        
        for col, date in date_columns:
            if task_start <= date <= task_end:
                cell = ws.cell(row=row, column=col)
                cell.fill = task_fill
                cell.border = border
    
    # Center align all cells
    for row in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=col-1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
    
    # Freeze panes
    ws.freeze_panes = 'B4'
    
    # Save workbook
    wb.save(filename) 